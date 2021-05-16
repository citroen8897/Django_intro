import openpyxl
import os
import shutil


def zip_list_excel(zip_name, zip_date, zip_quantity_docs, zip_link):
    """Формирование реестра архивов в формате .xlsx"""

    try:
        wb = openpyxl.load_workbook('./static/reestry/архивы_реестр.xlsx')
        sheet = wb['Sheet']

    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet['A1'] = 'Название архива'
        sheet['B1'] = 'Дата формирования'
        sheet['C1'] = 'Кол-во док-тов в архиве'
        sheet['D1'] = 'Ссылка'

    finally:
        sheet.append([zip_name, zip_date, zip_quantity_docs, zip_link])
        wb.save(filename='./static/reestry/архивы_реестр.xlsx')


def make_zip(queryset_list, zip_name):
    """Подготовка необходимых директорий, реестра документов,
    копирование документов, формирование архива, удаление
    вспомогательных директорий"""

    directory_main_name = zip_name
    parent_path = './static/'
    path = os.path.join(parent_path, directory_main_name)
    os.mkdir(path)

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet['A1'] = 'Регистрационный номер'
    sheet['B1'] = 'Дата создания'
    sheet['C1'] = 'Короткое описание'
    sheet['D1'] = 'Ссылка'

    directory_documents_name = 'Документи'
    parent_path = f'./static/{directory_main_name}'
    path = os.path.join(parent_path, directory_documents_name)
    os.mkdir(path)

    parent_path = f'./static/{directory_main_name}/Документи'
    for document in queryset_list:
        doc_reg_number = document.reg_number
        doc_name = document.title
        doc_file = document.main_file
        doc_date = document.create_date
        doc_content = document.comment
        doc_link = f'http://127.0.0.1:8000/api/document/{document.id}'

        sheet.append([doc_reg_number, doc_date, doc_content, doc_link])
        wb.save(filename=f'./static/{directory_main_name}/Реестр_документов.xlsx')

        directory_current_doc_name = doc_reg_number
        path = os.path.join(parent_path, directory_current_doc_name)
        os.mkdir(path)
        shutil.copyfile(f"{doc_file.file}",
                        f"./static/{directory_main_name}/Документи/{directory_current_doc_name}/{doc_file.name[9:]}")

    zip_name = f'./static/{directory_main_name}'
    directory_name = f'./static/{directory_main_name}'
    shutil.make_archive(zip_name, 'zip', directory_name)

    shutil.rmtree(directory_name)
